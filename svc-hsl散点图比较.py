import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os
"""本程序完成sig文件转换为xlsx文件"""
"coding: utf-8"
plt.style.use('ggplot')

# 判定是否可以转为数字


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass
    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
    return False


# file_dir  存储的路径
# data_list 二维数组
def write_xlsx(file_dir, data_list):
    workbook = openpyxl.Workbook()
    booksheet = workbook.active
    booksheet.title = "sheet1"
    row = 1
    for data in data_list:
        col = 1
        for ele in data:
            if is_number(ele):
                ele = float(ele)
            booksheet.cell(row, col, ele)
            col += 1
        row += 1
    workbook.save(file_dir)


# 返回np.array类型的数据
def read_sig_data(file_dir):
    if os.path.isdir(file_dir):
        return None
    f = open(file_dir)
    lines = f.readlines()
    data_list = []
    for line in lines:
        line = line.replace("\n", "")
        line_list = line.split("  ")  # 按照两个空格划分
        one_list = []
        for ele in line_list:
            one_list.append(ele)
        data_list.append(one_list)
    return data_list

# 读取xlsx中的文件


def read_xlsx_data(file_dir):
    if os.path.isdir(file_dir):
        return None
    wb = openpyxl.open(file_dir)
    sheet = wb.active
    test_data = []  # 创建一个空列表
    for row in range(1, sheet.max_row + 1):
        sub_data = []
        for column in range(1, sheet.max_column + 1):
            # Excel的第一行数据作为字典的key；
            sub_data.append(sheet.cell(row, column).value)
        test_data.append(sub_data)  # 将每行的数据循环加到列表中
    return test_data


def draw_scatter_plot(x, y, myplt, lab, mark):
    scatter, = myplt.scatter(x, y, s=20, c="#ff1212", marker='o', label=lab)
    myplt.xlim(500, 1000)
    # myplt.show()
    return scatter, myplt


def draw_plot(x, y, myplt, lab, mark):
    plot, = myplt.plot(x, y, label=lab, marker=mark)
    return plot, myplt


# 转换文件
if __name__ == "__main__":
    read_dir = "/Users/tianwenxin/Desktop/data/SVC_data_20210125/"
    file_list = ["000001_0000_R163_T165.sig", ]
    myplt = plt
    for file_name in file_list:
        data_list = read_sig_data(read_dir+file_name)
        data_list = data_list[25:]
        data_list = np.array(data_list, dtype=float)
        print("data_list shape: ", data_list.shape)
        data_list = np.transpose(data_list)
        scatter, myplt = draw_plot(
            data_list[0], data_list[3] / 100, myplt, "SVC", '.')
        myplt.xlabel("Wavelength(nm)")
        myplt.ylabel("Reflectance")

    read_dir2 = "/Users/tianwenxin/Desktop/result/HSL_ref/"
    file_list2 = ["木制品.xlsx"]
    for file_name in file_list2:
        data_list = read_xlsx_data(read_dir2+file_name)
        data_list = np.array(data_list, dtype=float)
        plot, myplt = draw_plot(data_list[0], data_list[3], myplt, "HSL", "p")

    # read_dir3 = "/Users/tianwenxin/Desktop/result/HSL_ref/"
    # file_list3 = ["金属样本40度.xlsx"]
    # for file_name in file_list3:
    #     data_list = read_xlsx_data(read_dir3+file_name)
    #     data_list3= np.array(data_list, dtype=float)
    #     plot, myplt = draw_plot(data_list[0], data_list[3], myplt, "HSL", "p")

    # 设置图像属性
    myplt.xlim(650, 900)
    myplt.ylim(0, 1)
    myplt.legend()
    myplt.show()

    print("转换完毕!")
