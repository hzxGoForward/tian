"""
该程序从原始文件中提取波形
读取新版本的dat文件
"""
import os
import numpy as np
import openpyxl
import matplotlib
import collections


# 从dat文件中，提取后一半数据的最大值
def readDat(dat_dir,excel_dir, cnt,):
    if os.path.exists(dat_dir) ==False:
        print("{0} does not exists".format(dat_dir))
        return
    f = open(dat_dir, 'r')
    lines = f.readlines()
    start = int(float(lines[2].replace('\n', '')))
    
    step = int(float(lines[4].replace('\n', '')))
    end = int(float(lines[3].replace('\n', ''))) + step
    lines = lines[9:-1]
    f.close()
    
    dict = collections.OrderedDict()
    band_list = []
    # 起始波段500，步长为10，截止770
    for band in range(start, end, step):
        band_list.append(band) # [500,510,520,...,770]
        dict[band] = []  
    print("正在处理中...")
    line_number = 0
    interval = 5000
    i = 0
    while i < cnt:
        for band in band_list:
            dict[band].append(lines[line_number: line_number + interval])
            line_number = line_number + interval
        i += 1
    mean_dict = collections.OrderedDict()
    for key in dict:
        sum = None
        for val in dict[key]:
            if sum is None:
                sum = np.array(val,dtype = int)
            else:
                sum += np.array(val,dtype = int)
        sum = sum/ len(dict[key])
        mean_dict[key] = sum
    
    print("写入结果中...")
    # 将结果写入excel
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    booksheet = workbook.create_sheet("data")
    maxsheet = workbook.create_sheet("max")
    col = 1
    for key in mean_dict:
        booksheet.cell(1, col, key)
        maxsheet.cell(1, col, key)
        maxsheet.cell(2, col, np.max(mean_dict[key]))
        maxsheet.cell(3, col, np.argmax(mean_dict[key])+1)
        for row in range(len(mean_dict[key])):
            booksheet.cell(row+2, col, mean_dict[key][row])
        col += 1
    workbook.save(excel_dir)

if __name__ == "__main__":
    read_dir = "/Users/tianwenxin/Desktop/data/20210124/0125-500-770data/" 
    save_dir = "/Users/tianwenxin/Desktop/result/20210124/"
    file_list = os.listdir(read_dir)
    cnt = 100
    for filename in file_list:
        if filename.endswith('.dat'):
            excel_dir = save_dir + filename.replace('.dat','.xlsx')
            dat_dir = read_dir + filename
            readDat(dat_dir,excel_dir, cnt)
            print('{} finished'.format(filename))
    print("totaly finished-------")
    
    