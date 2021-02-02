"""本程序完成sig文件转换为xlsx文件"""
"coding: utf-8"
import os
import openpyxl
import numpy as np

# 将单个.sig文件转换为xlsx文件
def convertsigToXlsx(openFileDir, saveFileDir):
    if os.path.isdir(openFileDir):
        return
    f = open(openFileDir)
    lines = f.readlines()
    lines = lines[25:]
    workbook = openpyxl.Workbook()
    booksheet = workbook.active
    booksheet.title = "sheet1"
    # 存储
    row = 1
    dt_list = []
    for line in lines:
        line = line.replace("\n","")
        dataList = line.split("  ")
        col = 1
        tmpData = []
        for data in dataList:
            if data.startswith(' ') or data == '':
                continue
            data = float(data)
            tmpData.append(data)
            booksheet.cell(row, col, data)
            col += 1
        if len(dt_list) == 0:
            dt_list.append(tmpData)
        elif tmpData[2] > dt_list[0][2]:
            dt_list.pop()
            dt_list.append(tmpData)
        row += 1
    print("正在写入: ", saveFileDir)
    workbook.save(saveFileDir)
    return dt_list[0]

def writeXlsx(file_dir,dataList):
    workbook = openpyxl.Workbook()
    booksheet = workbook.active
    booksheet.title = "sheet1"
    row = 1
    for data in dataList:
        col = 1
        for ele in data:
            booksheet.cell(row, col, float(ele))
            col += 1
        row += 1
    workbook.save(file_dir);

# 转换文件
if __name__ == "__main__":
    read_dir = "/Users/tianwenxin/Desktop/data/bochangdingbiao/HR1024i Data/"
    file_list = os.listdir(read_dir)
    save_dir = "/Users/tianwenxin/Desktop/result/bochangdingbiao/"
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    data_res = []
    for filename in file_list:
        newfilename = filename.replace(".sig",".xlsx")
        tmpData = convertsigToXlsx(read_dir + filename, save_dir + newfilename)
        tmpData.insert(0, newfilename.replace(".xlsx",''))
        data_res.append(tmpData)
        # data = np.array(data, dtype = float)
    writeXlsx(save_dir+"statics.xlsx", np.array(data_res))
    print("转换完毕!")