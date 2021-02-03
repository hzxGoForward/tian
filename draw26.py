import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os
plt.style.use('ggplot')

# 绘制26波段的图
# 2021年2月3日


def style():
    mark = [
        '.',
        ',',
        'o',
        'v',
        '^',
        '<',
        '>',
        '1',
        '2',
        '3',
        '4',
        '8',
        's',
        'p',
        'P',
        '*',
        'h',
        'H',
        '+',
        'x',
        'D',
        'd',
        '_', ]
    return mark


# 读取excel数据
def read_xlsx_data(file_dir):
    if os.path.isdir(file_dir):
        return None

    wb = openpyxl.load_workbook(file_dir)
    sheet = wb.active
    test_data = []  # 创建一个空列表
    for row in range(1, sheet.max_row + 1):
        sub_data = []
        for column in range(1, sheet.max_column + 1):
            # Excel的第一行数据作为字典的key；
            sub_data.append(sheet.cell(row, column).value)
        test_data.append(sub_data)  # 将每行的数据循环加到列表中
    return test_data


def draw_plot(x, y, myplt, lab, mark):
    plot, = myplt.plot(x, y, label=lab, marker=mark)
    return plot, myplt


if __name__ == '__main__':
    file_path = os.path.join(os.getcwd(), '100%.xlsx')
    data = read_xlsx_data(file_path)
    data = np.array(data, dtype=float)
    data = data.transpose()
    mark_list = style()
    myplt = plt
    band = 650
    for i in range(1, len(data)):
        plot, myplt = draw_plot(
            data[0], data[i], myplt, str(band), mark_list[(i-1) % (len(mark_list))])
        band += 10
    myplt.legend()
    myplt.xlabel("angle")
    myplt.ylabel("reflectance")
    myplt.show()
