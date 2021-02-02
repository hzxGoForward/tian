import numpy as np
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']
from scipy.optimize import curve_fit
from scipy import optimize
import math
import openpyxl

f0 = 0
def func1(x, k, n):
    a = np.cos(x*np.pi/180)
    b = np.cos(2*x*np.pi/180)
    n = (int)(abs(n))
    return (1-k)*a + k*pow(b, n)

# 估算kd和m参数
def estimate_kd_m(x, kd, m):
    x = x*np.pi/180
    cs = np.cos(x)
    mi = -pow(np.tan(x),2)/pow(m,2)
    up = pow(np.e, mi)/pow(cs, 5)
    global f0
    return f0*(kd*cs + (1-kd)*up)
 
 # 根据y和x算出f0
def correct_f0(x, y, kd, m):
    x = x*np.pi/180
    cs = np.cos(x)
    mi = -pow(np.tan(x),2)/pow(m,2)
    up = pow(np.e, mi)/pow(cs, 5)
    return y/(kd*cs + (1-kd)*up)


def get_threshold(x_list, kd, m):
    thresh_hold = x_list[-1]
    for angle in x_list:
        x = angle*np.pi/180
        cos_x = np.cos(x)
        mi = -pow(np.tan(x),2)/pow(m,2)
        up = pow(np.e, mi)/pow(cos_x, 5)
        res = (1-kd)*up
        if res < 0.1:
            thresh_hold = angle
            break
    return thresh_hold

def new_correct_fun(angle, kd, m, x_list, y):
    correct_list = []
    for i in range(0, len(x_list)):
        x = x_list[i]*np.pi/180
        cos_x = np.cos(x)
        if x_list[i] < angle:
            mi = -pow(np.tan(x),2)/pow(m,2)
            up = pow(np.e, mi)/pow(cos_x, 5)
            correct_list.append( (y[i] - y[0]*(1-kd)*up) / cos_x )
        else:
            correct_list.append( y[i]/cos_x )
    return correct_list
            


def draw_image(x, y, y_pred, y_f0, y_fpred0, title):
    y_f0 = np.array(y_f0)
    y_fpred0 = np.array(y_fpred0)
    print("f0: {}, y_fpred0:{}".format(f0, y_fpred0))
    line1, = plt.plot(x, y, 'r--' )
    line2, = plt.plot(x, y_pred, 'g--')
    line3, = plt.plot(x, y_f0, "y--")
    plt.legend([line1, line2, line3], ["original", "fit","jiaozheng"], loc = "upper right")
    plt.title(title)
    plt.show()

def fit(x, y, fun):
    bd = ([0, 0], [1, 0.6]) 
    popt, pcov = curve_fit(estimate_kd_m, x, y, bounds = bd)
    print("ks: {}, m: {}\n方差: {}".format(popt[0], popt[1], pcov))
    return popt

# 从excel中读取x和y
def read_data(dir):
    wb = openpyxl.load_workbook(dir)
    # 获取所有的工作表名
    names = wb.sheetnames
    dic = {}
    dic_band_list = {}
    for name in names:
        print("当前波段: {}".format(name))
        sheet = wb[name]
        y_list = []
        band_list = []
        for one_col_data in sheet.iter_cols():
            tmp = []
            first = 0
            for data in one_col_data:
                if first == 0:
                    band_list.append(data.value)
                    first += 1
                    continue;
                tmp.append(data.value)
            y_list.append(tmp)
        y_list = y_list[1:]
        dic[name] = y_list
        dic_band_list[name] = band_list
    wb.close()
    return (dic, dic_band_list)

def writeXlsx(fitDic, saveDir, title, band_list):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for key in fitDic:
        worksheet = workbook.create_sheet(key)
        row = 1
        col = 1
        for ce in title:
            worksheet.cell(row, col, ce)
            col += 1
        row += 1
        idx = 0
        for pair in fitDic[key]:
            band = band_list[idx]
            worksheet.cell(row, 1, band)
            worksheet.cell(row, 2, pair[0])
            worksheet.cell(row, 3, pair[1])
            row += 1
            idx += 1
    workbook.save(saveDir)

# 判定是否可以转为数字
def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass
    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
    return False

# file_dir  存储的路径
# data_list 二维数组
def write_xlsx(file_dir, data_list):
    workbook = openpyxl.Workbook()
    booksheet = workbook.active
    booksheet.title = "sheet1"
    row = 1
    for data in data_list:
        col = 1
        for ele in data:
            if is_number(ele):
                ele = float(ele)
            booksheet.cell(row, col, ele)
            col += 1
        row += 1
    workbook.save(file_dir)


def setf0(y):
    global f0
    f0 = y[0]
  

def fitAll():
    dir = "/Users/tianwenxin/Desktop/result/500-900HSL角度数据汇总.xlsx"
    # dir = "/home/hzx/Desktop/leaf_angle_all.xlsx"
    x = np.array([0, 10, 20, 30, 40, 50, 60, 70, 80])
    final_correct_list = [["name", "band"] + [0, 10, 20, 30, 40, 50, 60, 70, 80]]
    
    dic, dic_band = read_data(dir)
    fitDic = {}
    for key in dic:
        res = []
        idx = 0
        for y in dic[key]:
            start = dic_band[key][idx]
            setf0(y)
            data = fit(x, y, estimate_kd_m)
            alpha = get_threshold(x, data[0], data[1])
            kd = data[0]
            m = data[1]
            global f0
            title = "f0:{:.2f}-{}-{}-[threshold:{} kd:{:.2f} m:{:.2f} ]".format(f0, key, start, alpha, kd, m)
            print("------------当前:{}, f0: {}".format(title, f0))
            correct_list = new_correct_fun(alpha, kd, m, x, y)
            


            y_pred = np.array([estimate_kd_m(i, data[0], data[1]) for i in x])

            y_f0 = []
            y_fpred0 = []
            for i in range(len(x)):
                y_f0.append(correct_f0(x[i], y[i], data[0], data[1]))
                y_fpred0.append(correct_f0(x[i], y_pred[i], data[0], data[1]))

            line1, = plt.plot(x, y, 'r--', label = 'original' )
            line2, = plt.plot(x, y_pred, 'g--', label = 'fit')
            line4, = plt.plot(x, correct_list, 'b--', label = 'jiaozheng')
            plt.title(title)
            plt.legend()
            plt.show()


            correct_list.insert(0, key)
            correct_list.insert(1, start)
            final_correct_list.append(correct_list)

            idx += 1
            res.append(data)
        fitDic[key] = res
    rootDir = "/Users/tianwenxin/Desktop/result/"
    title = ["波段", "kd", "m"]
    saveDir = rootDir + "500-900HSL角度数据汇总-拟合.xlsx"
    writeXlsx(fitDic, saveDir, title, dic_band[key])


    
    save_dir = rootDir + "500-900HSL校正数据.xlsx"
    write_xlsx(save_dir, np.array(final_correct_list))

if __name__ == "__main__":
    fitAll()    



    